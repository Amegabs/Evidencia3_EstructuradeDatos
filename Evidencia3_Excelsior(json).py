import datetime
import csv
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sqlite3
from sqlite3 import Error
import warnings
warnings.filterwarnings("ignore",category=DeprecationWarning)


def iniciar_bd():
    """Funcion que crea la base de datos y las tablas"""
    try:
        with sqlite3.connect("coworking.db") as conexion:
            cursor = conexion.cursor()
            cursor.execute(
                """CREATE TABLE IF NOT EXISTS Clientes (id_cliente INTEGER PRIMARY KEY, nombre TEXT NOT NULL, apellido Text NOT NULL)"""
            )
            cursor.execute(
                """CREATE TABLE IF NOT EXISTS Salas (id_sala INTEGER PRIMARY KEY, nombre TEXT NOT NULL, cupo INTEGER NOT NULL)"""
            )
            cursor.execute(
                """CREATE TABLE IF NOT EXISTS Reservaciones (id_reservaciones INTEGER PRIMARY KEY, id_cliente INTEGER NOT NULL, id_sala INTEGER NOT NULL, fecha TEXT NOT NULL, turno TEXT NOT NULL, evento TEXT NOT NULL, FOREIGN KEY(id_cliente) REFERENCES Clientes(id_cliente), FOREIGN KEY(id_sala) REFERENCES Salas(id_sala))"""
            )
            print("TABLAS CREADAS EXITOSAMENTE.")
            main()
    except Error as e:
        print(e)
    except Exception as e:
        print(f"Error inesperado: {e}")


def pedir_nombres():
    """Funcion que pide el nombre"""
    while True:
        nombre = input("Nombre(s): ")
        if nombre == "":
            return None
        if nombre.replace(" ", "").isalpha():
            return nombre
        print("Error, ingresa el nombre correctamente. Solo letras.")


def pedir_apellidos():
    """Funcion que pide el apellido"""
    while True:
        apellido = input("Apellido(s): ")
        if apellido == "":
            return None
        if apellido.replace(" ", "").isalpha():
            return apellido
        print("Error, Ingresa el apellido correctamente. Solo letras.")


def pedir_cupo():
    """Funcion que pide el cupo de la sala"""
    while True:
        cupo = input("Cupo de la sala: ")
        if cupo == "":
            return None
        if cupo.isdigit():
            return cupo
        print("Error, Ingresa el número correctamente. Solo números.")


def registrar_reservacion():
    """Funcion que registrara una nueva reservacion en alguna sala disponible"""
    try:
        with sqlite3.connect("coworking.db") as conexion:
            cursor = conexion.cursor()
            cursor.execute("SELECT id_cliente, nombre, apellido FROM Clientes ORDER BY apellido, nombre")
            clientes = cursor.fetchall()
    except Error as e:
        print(e)
        return

    if not clientes:
        print(
            "No hay clientes registrados. Primero debe registrar un cliente para continuar."
        )
        return

    print("*"*75)
    print(f"**{'CLIENTES REGISTRADOS':^71}**")
    print("*"*75)
    print("{:<15} {:<30} {:<30} ".format("Clave Cliente", "Nombre(s)", "Apellido(s)"))
    print("*" * 75)
    for cliente in clientes:
        print("{:<15} {:<30} {:<30}".format(cliente[0], cliente[1], cliente[2]))
    else:
        print("*"*75)

    intentos_cliente = 0
    while True:
        entrada = input("Ingrese la Clave del Cliente: ").strip()
        print("")
        if entrada == "":
            intentos_cliente += 1
            if intentos_cliente >= 2:
                print("No ingresaste un cliente. Presiona ENTER para regresar al menu.")
                return
            else:
                print("No ingresaste un cliente. Intentalo de nuevo.")
                continue

        try:
            id_cliente = int(entrada)
        except ValueError:
            print("Debe ingresar un numero valido")
            continue
        if not any(cliente[0] == id_cliente for cliente in clientes):
            print("Cliente no encontrado. Intente de nuevo.")
            continue
        break

    try:
        with sqlite3.connect("coworking.db") as conexion:
            cursor = conexion.cursor()
            cursor.execute("SELECT id_sala, nombre, cupo FROM Salas")
            salas = cursor.fetchall()
    except Error as e:
        print(e)
        return

    if not salas:
        print("No hay salas registradas. Primero registre una sala para continuar.")
        return

    print("*"*65)
    print(f"**{'SALAS REGISTRADOS':^61}**")
    print("*"*65)
    print("{:<15} {:<30} {:<20} ".format("Clave Sala", "Nombre Sala", "Cupo"))
    print("*" * 65)
    for sala in salas:
        print("{:<15} {:<30} {:<20}".format(sala[0], sala[1], sala[2]))
    else:
        print("*"*65)

    intentos_sala = 0
    while True:
        entrada = input("Ingrese la Clave de la sala: ").strip()
        if entrada == "":
            intentos_sala += 1
            if intentos_sala >= 2:
                print("No ingresaste una sala. Presiona ENTER para regresar al menu.")
                return
            else:
                print("No ingresaste una sala. Intentalo de nuevo.")
                continue

        try:
            id_sala = int(entrada)
        except ValueError:
            intentos_sala += 1
            print("Debe ingresar un numero valido.")
            continue
        if not any(sala[0] == id_sala for sala in salas):
            print("Sala no encontrada. Intente de nuevo.")
            continue
        break

    intentos_fecha = 0
    while True:
        Fecha_str = input("Ingrese la fecha del evento (DD-MM-AAAA): ").strip()
        if Fecha_str == "":
            intentos_fecha += 1
            if intentos_fecha >= 2:
                print("No ingresaste una fecha. Presiona ENTER para regresar al menu.")
                return
            else:
                print("No ingresaste una fecha. Intentelo de nuevo.")
                continue

        try:
            Hoy = datetime.date.today()
            Fecha_evento = datetime.datetime.strptime(Fecha_str, "%d-%m-%Y").date()
            FechaAnticipada = (Fecha_evento - Hoy).days

            if FechaAnticipada < 2:
                intentos_fecha += 1
                if intentos_fecha >= 2:
                    print(
                        "La reservacion debe ser mayor a 2 dias de anticipacion. Intentalo nuevamente o ENTER para regresar al menu."
                    )
                    return
                else:
                    print(
                        "La reservacion debe ser mayor a 2 dias de anticipacion. Intentelo de nuevo."
                    )
                    continue

            if Fecha_evento.weekday() == 6:
                print("No se pueden realizar reservaciones los domingos.")
                lunes_siguiente = Fecha_evento + datetime.timedelta(days=1)
                print(f"Se propone automáticamente el lunes siguiente: {lunes_siguiente.strftime('%d-%m-%Y')}")
                respuesta = input("¿Aceptas esta fecha? (S/N): ").strip().upper()
                if respuesta == "S":
                    Fecha_evento = lunes_siguiente
                else:
                    print("Por favor, ingresa otra fecha que cumpla las condiciones.")

            break
        except ValueError:
            intentos_fecha += 1
            if intentos_fecha >= 2:
                print(
                    "Formato de fecha Incorrecto, use DD-MM-AAAA o ENTER para volver al menu."
                )
                return
            else:
                print(
                    "Formato de fecha incorrecto, use DD-MM-AAAA. Intentelo de nuevo."
                )
                continue

    intentos_turno = 0
    while True:
        turno = (
            input("Que turno desea? (Matutino/Vespertino/Nocturno): ").lower().strip()
        )

        if turno == "":
            intentos_turno += 1
            if intentos_turno >= 2:
                print(
                    "No ingresaste un turno. Intentalo nuevamente o ENTER para regresar al menu."
                )
                return
            else:
                print("No se ingreso un turno. Intentalo de nuevo.")
                continue

        if turno not in ["matutino", "vespertino", "nocturno"]:
            intentos_turno += 1
            if intentos_turno >= 2:
                print(
                    "Turno invalido. Intentalo nuevamente o ENTER para regresar al menu."
                )
                continue
            else:
                print("Turno inválido. Intentalo de nuevo.")
                return

        try:
            with sqlite3.connect("coworking.db") as conexion:
                cursor = conexion.cursor()
                cursor.execute(
                    """SELECT * FROM Reservaciones WHERE id_sala=? AND fecha=? AND turno=?""",
                    (id_sala, Fecha_evento.strftime("%d-%m-%Y"), turno),
                )
                ocupado = cursor.fetchone()
        except Error as e:
            print(e)
            return
        if ocupado:
            intentos_turno += 1
            if intentos_turno >= 2:
                print(
                    "Esta sala ya está reservada en esta fecha y turno. Intenta otro turno o ENTER para volver al menú."
                )
                return
            else:
                print(
                    "Esta sala ya esta reservada en esa fecha y turno. Intentelo de nuevo."
                )
                continue
        break
    
    intentos_evento = 0
    while True:
        nombre_evento = input("Ingrese el Nombre de su evento: ").strip()
        if nombre_evento == "":
            intentos_evento += 1
            if intentos_evento >= 2:
                print("No ingresaste un nombre de evento valido. Intentalo de nuevo o presiona ENTER para volver al menu.")
                return
            else:
                print("No ingresaste un nombre de evento valido. Intentalo de nuevo")
                continue
        break

    try:
        with sqlite3.connect("coworking.db") as conexion:
            cursor = conexion.cursor()
            cursor.execute(
                """
                INSERT INTO Reservaciones (id_cliente, id_sala, fecha, turno, evento)
                VALUES (?, ?, ?, ?, ?)
            """,
                (
                    id_cliente,
                    id_sala,
                    Fecha_evento,
                    turno,
                    nombre_evento,
                ),
            )
            conexion.commit()
            print("Reservación registrada exitosamente.")
    except Error as e:
        print(e)


def editar_reservacion():
    """Funcion que editara el nombre de la reservacion seleccionada por un rango de fechas"""
    try:
        fecha_inicio_str = input("Ingresa la fecha de inicio (DD-MM-AAAA): ").strip()
        fecha_fin_str = input("Ingresa la fecha de fin (DD-MM-AAAA): ").strip()

        fecha_inicio = datetime.datetime.strptime(fecha_inicio_str, "%d-%m-%Y").date()
        fecha_fin = datetime.datetime.strptime(fecha_fin_str, "%d-%m-%Y").date()

        fecha_inicio_iso = fecha_inicio.strftime("%Y-%m-%d")
        fecha_fin_iso = fecha_fin.strftime("%Y-%m-%d")

        with sqlite3.connect("coworking.db") as conexion:
            cursor = conexion.cursor()
            cursor.execute(
                """SELECT id_reservaciones, id_cliente, id_sala, fecha, turno, evento FROM Reservaciones WHERE fecha BETWEEN ? AND ?""",
                (fecha_inicio_iso, fecha_fin_iso),
            )
            reservaciones = cursor.fetchall()

        if not reservaciones:
            print("No se encontraron reservaciones en el rango indicado.")
            return

        print("*"*97)
        print(f"**{'RESERVACIONES ENCONTRADAS':^92}** ")
        print("*"*97)
        print("{:<20} {:<15} {:<15} {:<12} {:<10} {:<30}".format(
            "Clave Reservacion", "Clave cliente", "Clave sala", "Fecha", "Turno", "Nombre evento"))
        print("*" * 97)
        for reserva in reservaciones:
            print("{:<20} {:<15} {:<15} {:<12} {:<10} {:<30}".format(
                reserva[0], reserva[1], reserva[2], reserva[3], reserva[4], reserva[5]))
        else:
            print("*"*97)

        while True:
            try:
                clave_edit = int(input("\nIngrese la clave de la reservación que desea editar: ").strip())
                if clave_edit not in [r[0] for r in reservaciones]:
                    print("Clave inválida, intente de nuevo.")
                    continue
                break
            except ValueError:
                print("Debe ingresar un número válido.")

        nuevo_nombre = input("Ingrese el nuevo nombre del evento: ").strip()
        if not nuevo_nombre:
            print("El nombre del evento no puede quedar vacío.")
            return

        with sqlite3.connect("coworking.db") as conexion:
            cursor = conexion.cursor()
            cursor.execute(
                "UPDATE Reservaciones SET evento=? WHERE id_reservaciones=?",
                (nuevo_nombre, clave_edit)
            )
            conexion.commit()
            print("Nombre del evento actualizado con éxito.")

    except ValueError:
        print("Error en el formato de fechas, use DD-MM-AAAA.")
    except sqlite3.Error as e:
        print(f"Error en la base de datos: {e}")


def consultar_reservacion():
    """Funcion que consultara las reservaciones existentes para una fecha especifica"""
    try:
        while True:
            fecha_inicio_str = input("Ingrese la fecha inicial (DD-MM-AAAA): ").strip()
            if fecha_inicio_str == "":
                print("Consulta cancelada.")
                return
            try:
                fecha_inicio = datetime.datetime.strptime(fecha_inicio_str, "%d-%m-%Y").date()
                break
            except ValueError:
                print("Formato incorrecto. Use DD-MM-AAAA. Intente nuevamente.")

        while True:
            fecha_fin_str = input("Ingrese la fecha final (DD-MM-AAAA): ").strip()
            if fecha_fin_str == "":
                print("Consulta cancelada.")
                return
            try:
                fecha_fin = datetime.datetime.strptime(fecha_fin_str, "%d-%m-%Y").date()
                if fecha_fin < fecha_inicio:
                    print("La fecha final no puede ser menor que la fecha inicial.")
                    continue
                break
            except ValueError:
                print("Formato incorrecto. Use DD-MM-AAAA. Intente nuevamente.")

        with sqlite3.connect("coworking.db") as conexion:
            cursor = conexion.cursor()
            cursor.execute("SELECT * FROM Reservaciones")
            filas = cursor.fetchall()

    except Error as e:
        print(e)
        return

    if not filas:
        print("No hay reservaciones registradas.")
        return

    encabezados = ["Id Reserva", "Id Cliente", "Id Sala", "Fecha", "Turno", "Evento"]
    print("*"*74)
    print(f"**{'RESERVACIONES ENCONTRADAS':^70}**")
    print("*"*74)
    print("{:<12} {:<10} {:<10} {:<12} {:<10} {:<20}".format(*encabezados))
    print("*" * 74)

    encontrados = False
    for fila in filas:
        fecha_evento = datetime.datetime.strptime(fila[3], "%Y-%m-%d").date() if "-" in fila[3] else datetime.datetime.strptime(fila[3], "%d-%m-%Y").date()
        if fecha_inicio <= fecha_evento <= fecha_fin:
            print("{:<12} {:<10} {:<10} {:<12} {:<10} {:<20}".format(*fila))
            encontrados = True
    else:
        print("*"*74)

    if not encontrados:
        print(f"No hay reservaciones entre {fecha_inicio.strftime('%d-%m-%Y')} y {fecha_fin.strftime('%d-%m-%Y')}.")
        print("No hay reservaciones para esa fecha.")
    else:
        print("\n¿Deseas exportar los datos?")
        print("1. Exportar a CSV")
        print("2. Exportar a Excel")
        print("3. Exportar a JSON")
        print("4. No exportar")
        opcion = input("Selecciona una opción: ")

        if opcion == "1":
            exportar_csv(
                "reporte_reservaciones.csv", 
            )
        elif opcion == "2":
            exportar_excel()
        elif opcion == "3":
            exportar_json()
        else:
            print("No se exportaron los datos.")

def registrar_cliente():
    """Funcion que registrara a un nuevo cliente"""

    intentos_nombre = 0
    while True:
        nombre_cliente = pedir_nombres()
        if nombre_cliente:
            break
        intentos_nombre += 1
        if intentos_nombre >= 2:
            print("No ingresaste un nombre. Presiona ENTER para regresar al menú.")
            nombre_cliente = pedir_nombres()
            break
        else:
            print("No ingresaste un nombre. Intentalo nuevamente.")
    if not nombre_cliente:
        return

    intentos_apellido = 0
    while True:
        apellido_cliente = pedir_apellidos()
        if apellido_cliente:
            break
        intentos_apellido += 1
        if intentos_apellido >= 2:
            print("No ingresaste un apellido. Presiona ENTER para regresar al menú.")
            apellido_cliente = pedir_apellidos()
            break
        else:
            print("No ingresaste un apellido. Intentalo nuevamente.")
    if not apellido_cliente:
        return
    
    cliente = (nombre_cliente, apellido_cliente)

    try:
        with sqlite3.connect("coworking.db") as conexion:
            cursor = conexion.cursor()
            cursor.execute(
                "INSERT INTO Clientes (nombre, apellido) VALUES(?,?)", cliente
            )
            print("Cliente agregado exitosamente.")
    except Error as e:
        print(e)


def registrar_sala():
    """Funcion que registrara una nueva sala"""
    intentos_nombre = 0
    while True:
        nombre_sala = pedir_nombres()
        if nombre_sala:
            break
        intentos_nombre += 1
        if intentos_nombre >= 2:
            print(
                "No ingresaste un nombre de sala. Presiona ENTER para regresar al menu."
            )
            nombre_sala = pedir_nombres()
            break
        else:
            print("No ingresaste un nombre de sala. Intentalo nuevamente.")
    if not nombre_sala:
        return

    intentos_cupo = 0
    while True:
        cupo_sala = pedir_cupo()
        if cupo_sala:
            break
        intentos_cupo += 1
        if intentos_cupo >= 2:
            print(
                "No ingresaste un cupo de sala. Presiona ENTER para regresar al menu."
            )
            cupo_sala = pedir_cupo()
            break
        else:
            print("No ingresaste un cupo de sala. Intentalo nuevamente.")
    if not cupo_sala:
        return

    sala = (nombre_sala, cupo_sala)

    try:
        with sqlite3.connect("coworking.db") as conexion:
            cursor = conexion.cursor()
            cursor.execute("INSERT INTO Salas (nombre, cupo) VALUES(?,?)", sala)
            print("Sala creada exitosamente.")
    except Error as e:
        print(e)


def exportar_csv(nombre_archivo, reservaciones, salas, clientes):#QUE LOS DATOS SEAN DE LA BASE Y NO DE LOS DICCIONARIOS. MANTENIENDO EL FORMATEO CSV
    """Exporta las reservaciones a un archivo CSV"""
    try:
        if not reservaciones:
            print("No hay reservaciones")
            return

        with open(nombre_archivo, mode="w", encoding="utf-8", newline="") as file:
            writer = csv.writer(file)

            encabezados = ["Clave", "Cliente", "Sala", "Fecha", "Turno", "Evento"]
            writer.writerow(encabezados)
            writer.writerow(["-" * 7, "-" * 30, "-" * 20, "-" * 12, "-" * 10, "-" * 25])

            for clave, datos in reservaciones.items():
                clave_clientes = datos.get("Cliente")
                clave_salas = datos.get("Sala")

                nombre_cliente = (
                    f"{clientes[clave_clientes][0]} {clientes[clave_clientes][1]}"
                    if clave_clientes in clientes
                    else "Desconocido"
                )
                nombre_salas = (
                    salas[clave_salas][0] if clave_salas in salas else "Desconocido"
                )

                fecha_str = datos.get("Fecha").strftime("%d-%m-%Y")

                fila = [
                    clave,
                    nombre_cliente.ljust(30),
                    nombre_salas.ljust(20),
                    fecha_str,
                    datos.get("Turno"),
                    datos.get("Evento").ljust(25),
                ]
                writer.writerow(fila)

        print(f"Archivo {nombre_archivo} creado correctamente")

    except Exception as e:
        print(f"Error al crear el archivo: {e}")


def exportar_excel(nombreArchivo="DATOSEXCEL.xlsx"):#me when:
    """Funcion que exporta las reservaciones a un archivo Excel"""
    try:
        with sqlite3.connect("coworking.db") as conexion:
            Cursor = conexion.cursor()

            Cursor.execute("""
                SELECT 
                    R.id_reservaciones,
                    C.nombre || ' ' || C.apellido AS cliente,
                    S.nombre AS sala,
                    R.fecha,
                    R.turno,
                    R.evento
                FROM Reservaciones R
                JOIN Clientes C ON R.id_cliente = C.id_cliente
                JOIN Salas S ON R.id_sala = S.id_sala
                ORDER BY R.fecha;
            """)
            datos = Cursor.fetchall()

    except sqlite3.Error as e:
        print("Error al conectar con la base de datos:", e)
        return

    if not datos:
        print("No hay reservaciones en la base de datos para exportar.")
        return
    

    wb = Workbook()
    ws = wb.active
    ws.title = "Reservaciones"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    cell = ws.cell(row=1, column=1)
    cell.value = "REPORTE DE RESERVACIONES"
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal="center")

    ws.append([])

    encabezados = ["Clave Reservacion", "Cliente", "Sala", "Fecha", "Turno", "Evento"]  
    negrita = Font(bold=True)
    borde_grueso = Border(bottom=Side(style="thick"))
    alineado = Alignment(horizontal="center")

    for col, encabezado in enumerate(encabezados, start=1):
        celda = ws.cell(row=3, column=col, value=encabezado)
        celda.font = negrita
        celda.alignment = alineado
        celda.border = borde_grueso   

    for fila, dato in enumerate(datos, start=4):
        id_reservacion, cliente, sala, fecha, turno, evento = dato


        try:
            fecha = datetime.strptime(fecha, "%Y-%m-%d").strftime("%d-%m-%Y")
        except:
            pass

        ws.cell(row=fila, column=1, value=id_reservacion)
        ws.cell(row=fila, column=2, value=cliente)
        ws.cell(row=fila, column=3, value=sala)
        ws.cell(row=fila, column=4, value=fecha)
        ws.cell(row=fila, column=5, value=turno.capitalize())
        ws.cell(row=fila, column=6, value=evento)

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")

    for i, col_cells in enumerate(
        ws.iter_cols(min_row=3, max_row=ws.max_row, min_col=1, max_col=6), start=1
    ):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
        ws.column_dimensions[get_column_letter(i)].width = max_length + 2

    wb.save(nombreArchivo)
    print(f"Archivo '{nombreArchivo}' exportado correctamente.")



def exportar_json(nombre_archivo="DatosJSON.json"):
    """Exporta las reservaciones de la base de datos a un archivo JSON."""
    try:
        with sqlite3.connect("coworking.db") as conexion:
            cursor = conexion.cursor()
            cursor.execute("""
                SELECT 
                    R.id_reservaciones,
                    C.nombre || ' ' || C.apellido AS cliente,
                    S.nombre AS sala,
                    R.fecha,
                    R.turno,
                    R.evento
                FROM Reservaciones R
                JOIN Clientes C ON R.id_cliente = C.id_cliente
                JOIN Salas S ON R.id_sala = S.id_sala
                ORDER BY R.fecha;
            """)
            datos = cursor.fetchall()
    except sqlite3.Error as e:
        print(f"Error al conectar con la base de datos: {e}")
        return

    if not datos:
        print("No hay reservaciones para exportar a JSON.")
        return

    lista_datos = []
    for id_reservacion, cliente, sala, fecha, turno, evento in datos:
        try:
            if "-" in fecha:
                fecha = datetime.datetime.strptime(fecha, "%Y-%m-%d").strftime("%d-%m-%Y")
        except Exception:
            pass

        lista_datos.append({
            "Clave Reservacion": id_reservacion,
            "Cliente": cliente,
            "Sala": sala,
            "Fecha": fecha,
            "Turno": turno,
            "Evento": evento
        })

    try:
        with open(nombre_archivo, "w", encoding="utf-8") as file:
            json.dump(lista_datos, file, ensure_ascii=False, indent=4)
        print(f"Archivo '{nombre_archivo}' exportado correctamente.")
    except Exception as e:
        print(f"Error al exportar a JSON: {e}")


def main():
    while True:
        print("\nMENU PRINCIPAL.")
        print("Opciones disponibles: ")
        print("1. Registrar nueva reservacion.")
        print("2. Editar nombre de reservacion.")
        print("3. Consultar reservaciones.")
        print("4. Registrar nuevo cliente.")
        print("5. Registrar nueva sala.")
        print("6. Salir.\n")
        opcion = input("Selecciona la opcion que necesites (1-6): ")
        if opcion.isdigit():
            if opcion == "1":
                registrar_reservacion()
            elif opcion == "2":
                editar_reservacion()
            elif opcion == "3":
                consultar_reservacion()
            elif opcion == "4":
                registrar_cliente()
            elif opcion == "5":
                registrar_sala()
            elif opcion == "6":
                salir = input("¿Desea salir realmente? s/n: ").lower()
                if salir == "s":
                    print("Saliendo del programa...")
                    break
            else:
                print("Error, ingrese una opción valida")
        else:
            print("Error, ingrese una opción valida")


if __name__ == "__main__":
    iniciar_bd()
